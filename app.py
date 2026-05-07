import json
import re
import secrets

from datetime import datetime, timedelta

from flask import (
    Flask,
    request,
    redirect,
    url_for,
    render_template,
    send_file,
    session,
    flash
)

from werkzeug.security import generate_password_hash, check_password_hash

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from config import SECRET_KEY, WEBHOOK_TOKEN
from db import get_connection

app = Flask(__name__)
app.secret_key = SECRET_KEY

# =========================================================
# AUTH HELPERS
# =========================================================



def now_text():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def create_account_token(user_id, token_type, hours_valid=48):
    token = secrets.token_urlsafe(48)

    expires_at = (
        datetime.now() + timedelta(hours=hours_valid)
    ).strftime("%Y-%m-%d %H:%M:%S")

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
    INSERT INTO account_tokens (
        user_id,
        token,
        token_type,
        expires_at
    )
    VALUES (?, ?, ?, ?)
    """, (
        user_id,
        token,
        token_type,
        expires_at
    ))

    conn.commit()
    conn.close()

    return token


def get_valid_token(token, token_type):
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT
        account_tokens.*,
        users.email,
        users.full_name,
        users.is_active
    FROM account_tokens
    JOIN users
    ON users.id = account_tokens.user_id
    WHERE account_tokens.token = ?
    AND account_tokens.token_type = ?
    AND account_tokens.used_at IS NULL
    AND account_tokens.expires_at > CURRENT_TIMESTAMP
    LIMIT 1
    """, (
        token,
        token_type
    ))

    token_data = cursor.fetchone()

    conn.close()

    return token_data


def mark_token_used(token_id):
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
    UPDATE account_tokens
    SET used_at = ?
    WHERE id = ?
    """, (
        now_text(),
        token_id
    ))

    conn.commit()
    conn.close()



def login_required(view):
    def wrapped_view(*args, **kwargs):

        if not session.get("user_id"):
            return redirect(url_for("login"))

        return view(*args, **kwargs)

    wrapped_view.__name__ = view.__name__

    return wrapped_view


def extract_event_dates(value):
    raw_value = (value or "").strip()

    if not raw_value:
        return "", "", ""

    normalized = raw_value.replace("–", "-")
    normalized = normalized.replace("—", "-")

    patterns = [

        # 19.07-1.08.2026
        r"(\d{1,2})\.(\d{1,2})\s*-\s*(\d{1,2})\.(\d{1,2})\.(\d{4})",

        # 2-15.08.2026
        r"(\d{1,2})\s*-\s*(\d{1,2})\.(\d{1,2})\.(\d{4})",

        # 19.07.2026-01.08.2026
        r"(\d{1,2})\.(\d{1,2})\.(\d{4})\s*-\s*(\d{1,2})\.(\d{1,2})\.(\d{4})"
    ]

    for index, pattern in enumerate(patterns):

        match = re.search(pattern, normalized)

        if not match:
            continue

        # 19.07-1.08.2026
        if index == 0:

            start_day = match.group(1).zfill(2)
            start_month = match.group(2).zfill(2)

            end_day = match.group(3).zfill(2)
            end_month = match.group(4).zfill(2)

            year = match.group(5)

        # 2-15.08.2026
        elif index == 1:

            start_day = match.group(1).zfill(2)

            end_day = match.group(2).zfill(2)
            end_month = match.group(3).zfill(2)

            year = match.group(4)

            start_month = end_month

        # pełne daty
        else:

            start_day = match.group(1).zfill(2)
            start_month = match.group(2).zfill(2)

            year = match.group(3)

            end_day = match.group(4).zfill(2)
            end_month = match.group(5).zfill(2)

        date_from = f"{year}-{start_month}-{start_day}"
        date_to = f"{year}-{end_month}-{end_day}"

        date_label = (
            f"{start_day}.{start_month}"
            f"-"
            f"{end_day}.{end_month}.{year}"
        )

        return date_from, date_to, date_label

    return "", "", ""


def get_field(data, *names):
    for name in names:
        value = data.get(name)

        if value:
            return str(value).strip()

    return ""


@app.route("/")
def index():

    if session.get("user_id"):
        return redirect(url_for("leads"))

    return redirect(url_for("login"))




@app.route("/leads")
@login_required
def leads():
    selected_range = (request.args.get("range") or "").strip()

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT DISTINCT
        event_date_label,
        event_date_from
    FROM leads
    WHERE event_date_label IS NOT NULL
    AND event_date_label != ''
    ORDER BY event_date_from ASC
    """)

    event_ranges = cursor.fetchall()

    if selected_range:
        cursor.execute("""
        SELECT *
        FROM leads
        WHERE event_date_label = ?
        ORDER BY created_at DESC, id DESC
        """, (selected_range,))
    else:
        cursor.execute("""
        SELECT *
        FROM leads
        ORDER BY created_at DESC, id DESC
        """)

    leads_data = cursor.fetchall()

    conn.close()

    return render_template(
        "leads.html",
        leads=leads_data,
        event_ranges=event_ranges,
        selected_range=selected_range
    )


@app.route("/api/leads", methods=["POST"])
def receive_lead():

    print("===================================")
    print("NEW WEBHOOK REQUEST")
    print("===================================")

    print("HEADERS:")
    print(dict(request.headers))

    print("FORM:")
    print(request.form.to_dict())

    print("JSON:")
    print(request.get_json(silent=True))

    token = request.headers.get("X-WEBHOOK-TOKEN")

    if token != WEBHOOK_TOKEN:
        return {
            "ok": False,
            "error": "Unauthorized"
        }, 401

    data = request.get_json(silent=True)

    if not data:
        data = request.form.to_dict()

    source_url = (
        get_field(
            data,
            "source_url",
            "Source URL",
            "URL"
        )
        or request.headers.get("Referer")
        or ""
    ).strip()

    event_label = get_field(
        data,
        "Wybierz imprezę",
        "event_label",
        "event_name",
        "trip"
    )

    event_date_from, event_date_to, event_date_label = (
        extract_event_dates(event_label)
    )

    participant_full_name = get_field(
        data,
        "Imię i nazwisko uczestnika",
        "participant_full_name",
        "full_name"
    )

    birth_date = get_field(
        data,
        "Data urodzenia",
        "birth_date"
    )

    pesel = get_field(
        data,
        "numer PESEL",
        "PESEL",
        "pesel"
    )

    participant_phone = get_field(
        data,
        "Numer telefonu uczestnika",
        "participant_phone",
        "phone"
    )

    address = get_field(
        data,
        "Adres zamieszkania",
        "address"
    )

    city_postal_code = get_field(
        data,
        "Miasto i kod pocztowy",
        "city_postal_code"
    )

    guardian_full_name = get_field(
        data,
        "Imię i nazwisko opiekuna",
        "guardian_full_name"
    )

    guardian_phone = get_field(
        data,
        "Numer telefonu opiekuna",
        "guardian_phone"
    )

    guardian_email = get_field(
        data,
        "Adres e-mail opiekuna",
        "guardian_email",
        "email"
    )

    referrer_full_name = get_field(
        data,
        "Imię i nazwisko osoby polecającej",
        "referrer_full_name"
    )

    notes = get_field(
        data,
        "Uwagi",
        "notes",
        "message"
    )

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
    INSERT INTO leads (
        status,

        source_url,

        event_label,
        event_date_from,
        event_date_to,
        event_date_label,

        participant_full_name,
        birth_date,
        pesel,

        participant_phone,

        address,
        city_postal_code,

        guardian_full_name,
        guardian_phone,
        guardian_email,

        referrer_full_name,

        notes,

        raw_payload
    )
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        "new",

        source_url,

        event_label,
        event_date_from,
        event_date_to,
        event_date_label,

        participant_full_name,
        birth_date,
        pesel,

        participant_phone,

        address,
        city_postal_code,

        guardian_full_name,
        guardian_phone,
        guardian_email,

        referrer_full_name,

        notes,

        json.dumps(data, ensure_ascii=False)
    ))

    conn.commit()
    conn.close()

    return {
        "ok": True,
        "message": "Lead saved",
        "event_date_label": event_date_label
    }


@app.route("/test-lead")
@login_required
def test_lead():
    data = {
        "Wybierz imprezę": "Mazurski Szlak II Turnus (2-15.08.2026)",
        "Imię i nazwisko uczestnika": "Jan Kowalski",
        "Data urodzenia": "01.01.2010",
        "numer PESEL": "12345678901",
        "Numer telefonu uczestnika": "600700800",
        "Adres zamieszkania": "Testowa 1",
        "Miasto i kod pocztowy": "00-000 Warszawa",
        "Imię i nazwisko opiekuna": "Anna Kowalska",
        "Numer telefonu opiekuna": "600111222",
        "Adres e-mail opiekuna": "anna@test.pl",
        "Imię i nazwisko osoby polecającej": "",
        "Uwagi": "Test z Flask",
        "source_url": "https://podwiatr.org/812-2/"
    }

    event_label = data["Wybierz imprezę"]
    event_date_from, event_date_to, event_date_label = extract_event_dates(event_label)

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
    INSERT INTO leads (
        status,
        source_url,
        event_label,
        event_date_from,
        event_date_to,
        event_date_label,
        participant_full_name,
        birth_date,
        pesel,
        participant_phone,
        address,
        city_postal_code,
        guardian_full_name,
        guardian_phone,
        guardian_email,
        referrer_full_name,
        notes,
        raw_payload
    )
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        "new",
        data["source_url"],
        event_label,
        event_date_from,
        event_date_to,
        event_date_label,
        data["Imię i nazwisko uczestnika"],
        data["Data urodzenia"],
        data["numer PESEL"],
        data["Numer telefonu uczestnika"],
        data["Adres zamieszkania"],
        data["Miasto i kod pocztowy"],
        data["Imię i nazwisko opiekuna"],
        data["Numer telefonu opiekuna"],
        data["Adres e-mail opiekuna"],
        data["Imię i nazwisko osoby polecającej"],
        data["Uwagi"],
        json.dumps(data, ensure_ascii=False)
    ))

    conn.commit()
    conn.close()

    return redirect(url_for("leads"))


@app.route("/export/excel")
@login_required
def export_excel():
    selected_range = (request.args.get("range") or "").strip()

    conn = get_connection()
    cursor = conn.cursor()

    if selected_range:
        cursor.execute("""
        SELECT *
        FROM leads
        WHERE event_date_label = ?
        ORDER BY created_at ASC, id ASC
        """, (selected_range,))
    else:
        cursor.execute("""
        SELECT *
        FROM leads
        ORDER BY created_at ASC, id ASC
        """)

    leads_data = cursor.fetchall()
    conn.close()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Zgloszenia"

    headers = [
        "Data zgłoszenia",
        "Turnus",
        "Zakres dat",
        "Uczestnik",
        "Data urodzenia",
        "PESEL",
        "Telefon uczestnika",
        "Adres",
        "Miasto i kod pocztowy",
        "Opiekun",
        "Telefon opiekuna",
        "E-mail opiekuna",
        "Osoba polecająca",
        "Uwagi",
        "Źródło"
    ]

    sheet.append(headers)

    for lead in leads_data:
        sheet.append([
            lead["created_at"],
            lead["event_label"],
            lead["event_date_label"],
            lead["participant_full_name"],
            lead["birth_date"],
            lead["pesel"],
            lead["participant_phone"],
            lead["address"],
            lead["city_postal_code"],
            lead["guardian_full_name"],
            lead["guardian_phone"],
            lead["guardian_email"],
            lead["referrer_full_name"],
            lead["notes"],
            lead["source_url"]
        ])

    header_fill = PatternFill(
        start_color="F3F4F6",
        end_color="F3F4F6",
        fill_type="solid"
    )

    header_font = Font(
        color="111111",
        bold=True
    )

    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD")
    )

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(
            vertical="center",
            horizontal="left",
            wrap_text=True
        )

    column_widths = {
        "A": 20,
        "B": 42,
        "C": 20,
        "D": 28,
        "E": 16,
        "F": 18,
        "G": 20,
        "H": 34,
        "I": 26,
        "J": 28,
        "K": 20,
        "L": 30,
        "M": 26,
        "N": 42,
        "O": 42
    }

    for column, width in column_widths.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(
                vertical="top",
                horizontal="left",
                wrap_text=True
            )

        sheet.row_dimensions[row[0].row].height = 34

    sheet.freeze_panes = "A2"

    filename_suffix = (
        selected_range.replace(".", "-").replace("/", "-")
        if selected_range
        else "wszystkie"
    )

    filename = (
        f"zgloszenia_{filename_suffix}_"
        f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    filepath = f"exports/generated/{filename}"

    workbook.save(filepath)

    return send_file(
        filepath,
        as_attachment=True,
        download_name=filename
    )


@app.route("/leads/<int:lead_id>/delete", methods=["POST"])
@login_required
def delete_lead(lead_id):
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
    DELETE FROM leads
    WHERE id = ?
    """, (lead_id,))

    conn.commit()
    conn.close()

    return redirect(url_for("leads"))


# =========================================================
# AUTH ROUTES
# =========================================================

@app.route("/login", methods=["GET", "POST"])
def login():

    if session.get("user_id"):
        return redirect(url_for("leads"))

    if request.method == "POST":

        email = (request.form.get("email") or "").strip().lower()
        password = request.form.get("password") or ""

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
        SELECT *
        FROM users
        WHERE lower(email) = ?
        AND is_active = 1
        LIMIT 1
        """, (email,))

        user = cursor.fetchone()

        if not user or not user["password_hash"]:
            conn.close()
            flash("Nieprawidłowy email lub hasło.", "error")
            return redirect(url_for("login"))

        password_ok = check_password_hash(
            user["password_hash"],
            password
        )

        if not password_ok:
            conn.close()
            flash("Nieprawidłowy email lub hasło.", "error")
            return redirect(url_for("login"))

        cursor.execute("""
        UPDATE users
        SET last_login_at = ?
        WHERE id = ?
        """, (
            now_text(),
            user["id"]
        ))

        conn.commit()
        conn.close()

        session["user_id"] = user["id"]
        session["user_name"] = user["full_name"]
        session["user_email"] = user["email"]
        session["user_role"] = user["role"]

        return redirect(url_for("leads"))

    return render_template("login.html")


@app.route("/logout")
def logout():

    session.clear()

    return redirect(url_for("login"))


@app.route("/admin/users/create", methods=["GET", "POST"])
@login_required
def create_user():

    if request.method == "POST":

        full_name = (request.form.get("full_name") or "").strip()
        email = (request.form.get("email") or "").strip().lower()

        if not full_name or not email:
            flash("Uzupełnij imię i nazwisko oraz email.", "error")
            return redirect(url_for("create_user"))

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
        SELECT id
        FROM users
        WHERE lower(email) = ?
        LIMIT 1
        """, (email,))

        existing_user = cursor.fetchone()

        if existing_user:
            conn.close()
            flash("Użytkownik z takim adresem email już istnieje.", "error")
            return redirect(url_for("create_user"))

        cursor.execute("""
        INSERT INTO users (
            full_name,
            email,
            role,
            is_active
        )
        VALUES (?, ?, ?, ?)
        """, (
            full_name,
            email,
            "client",
            0
        ))

        user_id = cursor.lastrowid

        conn.commit()
        conn.close()

        token = create_account_token(
            user_id=user_id,
            token_type="activation",
            hours_valid=72
        )

        activation_link = url_for(
            "activate_account",
            token=token,
            _external=True
        )

        return render_template(
            "user_created.html",
            activation_link=activation_link,
            email=email
        )

    return render_template("create_user.html")


@app.route("/activate/<token>", methods=["GET", "POST"])
def activate_account(token):

    token_data = get_valid_token(
        token=token,
        token_type="activation"
    )

    if not token_data:
        flash("Link aktywacyjny jest nieprawidłowy albo wygasł.", "error")
        return redirect(url_for("login"))

    if request.method == "POST":

        password = request.form.get("password") or ""
        password_confirm = request.form.get("password_confirm") or ""

        if len(password) < 8:
            flash("Hasło musi mieć minimum 8 znaków.", "error")
            return redirect(url_for("activate_account", token=token))

        if password != password_confirm:
            flash("Hasła nie są takie same.", "error")
            return redirect(url_for("activate_account", token=token))

        password_hash = generate_password_hash(password)

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
        UPDATE users
        SET
            password_hash = ?,
            is_active = 1,
            activated_at = ?,
            updated_at = ?
        WHERE id = ?
        """, (
            password_hash,
            now_text(),
            now_text(),
            token_data["user_id"]
        ))

        conn.commit()
        conn.close()

        mark_token_used(token_data["id"])

        flash("Konto zostało aktywowane. Możesz się zalogować.", "success")
        return redirect(url_for("login"))

    return render_template(
        "activate_account.html",
        token=token,
        email=token_data["email"]
    )


@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():

    if request.method == "POST":

        email = (request.form.get("email") or "").strip().lower()

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
        SELECT *
        FROM users
        WHERE lower(email) = ?
        AND is_active = 1
        LIMIT 1
        """, (email,))

        user = cursor.fetchone()

        conn.close()

        if user:
            token = create_account_token(
                user_id=user["id"],
                token_type="password_reset",
                hours_valid=2
            )

            reset_link = url_for(
                "reset_password",
                token=token,
                _external=True
            )

            return render_template(
                "password_reset_created.html",
                reset_link=reset_link,
                email=email
            )

        flash("Jeżeli konto istnieje, link resetujący został przygotowany.", "success")
        return redirect(url_for("login"))

    return render_template("forgot_password.html")


@app.route("/reset-password/<token>", methods=["GET", "POST"])
def reset_password(token):

    token_data = get_valid_token(
        token=token,
        token_type="password_reset"
    )

    if not token_data:
        flash("Link resetujący jest nieprawidłowy albo wygasł.", "error")
        return redirect(url_for("login"))

    if request.method == "POST":

        password = request.form.get("password") or ""
        password_confirm = request.form.get("password_confirm") or ""

        if len(password) < 8:
            flash("Hasło musi mieć minimum 8 znaków.", "error")
            return redirect(url_for("reset_password", token=token))

        if password != password_confirm:
            flash("Hasła nie są takie same.", "error")
            return redirect(url_for("reset_password", token=token))

        password_hash = generate_password_hash(password)

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
        UPDATE users
        SET
            password_hash = ?,
            updated_at = ?
        WHERE id = ?
        """, (
            password_hash,
            now_text(),
            token_data["user_id"]
        ))

        conn.commit()
        conn.close()

        mark_token_used(token_data["id"])

        flash("Hasło zostało zmienione. Możesz się zalogować.", "success")
        return redirect(url_for("login"))

    return render_template(
        "reset_password.html",
        token=token,
        email=token_data["email"]
    )


@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():

    if request.method == "POST":

        current_password = request.form.get("current_password") or ""
        new_password = request.form.get("new_password") or ""
        new_password_confirm = request.form.get("new_password_confirm") or ""

        conn = get_connection()
        cursor = conn.cursor()

        cursor.execute("""
        SELECT *
        FROM users
        WHERE id = ?
        LIMIT 1
        """, (session["user_id"],))

        user = cursor.fetchone()

        if not user:
            conn.close()
            session.clear()
            return redirect(url_for("login"))

        password_ok = check_password_hash(
            user["password_hash"],
            current_password
        )

        if not password_ok:
            conn.close()
            flash("Obecne hasło jest nieprawidłowe.", "error")
            return redirect(url_for("change_password"))

        if len(new_password) < 8:
            conn.close()
            flash("Nowe hasło musi mieć minimum 8 znaków.", "error")
            return redirect(url_for("change_password"))

        if new_password != new_password_confirm:
            conn.close()
            flash("Nowe hasła nie są takie same.", "error")
            return redirect(url_for("change_password"))

        password_hash = generate_password_hash(new_password)

        cursor.execute("""
        UPDATE users
        SET
            password_hash = ?,
            updated_at = ?
        WHERE id = ?
        """, (
            password_hash,
            now_text(),
            user["id"]
        ))

        conn.commit()
        conn.close()

        flash("Hasło zostało zmienione.", "success")
        return redirect(url_for("leads"))

    return render_template("change_password.html")



if __name__ == "__main__":
    app.run(debug=True)